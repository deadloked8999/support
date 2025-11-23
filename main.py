import re
import uuid
import os
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, LabeledPrice
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ConversationHandler,
    PreCheckoutQueryHandler,
    filters,
    ContextTypes,
    JobQueue,
    ApplicationHandlerStop,
)
from database import (
    init_database,
    add_purchase,
    add_activation,
    update_activation_receipt,
    update_activation_kit,
    update_activation_serial_number,
    update_activation_serial_photo,
    update_activation_box_serial_number,
    update_activation_box_serial_photo,
    get_all_purchases,
    get_all_activations,
    get_statistics,
    mark_service_provided,
    get_activations_for_subscription_reminders,
    update_last_reminder_day,
    update_activation_email_password,
    get_activation_by_id,
    get_pending_activations,
    get_processed_activations,
    find_activation_by_request_number,
    find_purchase_by_request_number,
    delete_activation,
    delete_purchase,
    toggle_service_provided,
)
from config import BOT_TOKEN, ACTIVATION_PRICE, ACTIVATION_PRICE_TON, PAYMENT_PHONE, PROVIDER_TOKEN, ADMIN_IDS, ADMIN_PASSWORD, SERIAL_NUMBER_EXAMPLE


WAITING_PHONE_PURCHASE, WAITING_NAME_PURCHASE = range(2)
WAITING_PHONE_ACTIVATE, WAITING_NAME_ACTIVATE, WAITING_SERIAL, WAITING_SERIAL_PHOTO, WAITING_BOX_SERIAL, WAITING_BOX_SERIAL_PHOTO = range(5, 11)
WAITING_ADMIN_PASSWORD = 15
WAITING_ADMIN_SELECT_ACTIVATION, WAITING_ADMIN_EMAIL, WAITING_ADMIN_PASSWORD_FIELD = 16, 17, 18
WAITING_ADMIN_SEARCH = 19
WAITING_ADMIN_DELETE_CONFIRM = 20


def normalize_phone(phone):
    """Валидация и нормализация номера телефона.
    Требования: строго 11 цифр, начинается с +7 или 8.
    Возвращает нормализованный номер или None в случае ошибки.
    """
    phone = phone.strip()
    
    # Убираем все пробелы, дефисы, скобки для проверки
    phone_clean = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    
    # Проверяем, что остались только цифры (или +7 в начале)
    if phone_clean.startswith('+7'):
        phone_clean = phone_clean[2:]  # Убираем +7 для проверки
    
    # Проверяем, что все символы - цифры и длина равна 10 (после удаления +7) или 11 (если начинается с 8)
    if not phone_clean.isdigit():
        return None  # Есть буквы или другие символы
    
    # Обрабатываем разные варианты начала
    original_phone = phone.strip().replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    
    if original_phone.startswith('+7') and len(original_phone) == 12:  # +7 и 10 цифр = 12 символов
        return '+7' + original_phone[2:]
    elif original_phone.startswith('8') and len(original_phone) == 11:  # 8 и 10 цифр = 11 символов
        return '+7' + original_phone[1:]
    elif original_phone.startswith('7') and len(original_phone) == 11:  # 7 и 10 цифр = 11 символов
        return '+7' + original_phone[1:]
    
    return None  # Неправильный формат


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print(f"DEBUG: start command received from user {update.effective_user.id}")
    try:
        # Очищаем все состояния ConversationHandler для этого пользователя
        context.user_data.clear()
        print(f"DEBUG: user_data cleared for user {update.effective_user.id}")
        
        welcome_text = (
            "Добро пожаловать! 👋\n\n"
            "Это техподдержка по активации терминалов Starlink. "
            "Я помогу вам купить терминал или активировать уже имеющееся устройство.\n\n"
            "Выберите нужное действие:"
        )
        
        keyboard = [
            [InlineKeyboardButton("🛒 Купить терминал", callback_data="buy")],
            [InlineKeyboardButton("⚙️ Активировать", callback_data="activate")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(welcome_text, reply_markup=reply_markup)
        print(f"DEBUG: start message sent to user {update.effective_user.id}")
        
        # Останавливаем дальнейшую обработку
        raise ApplicationHandlerStop()
    except ApplicationHandlerStop:
        raise
    except Exception as e:
        print(f"Ошибка в start: {e}")
        import traceback
        traceback.print_exc()
        raise ApplicationHandlerStop()


async def button_callback_buy(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    # Очищаем данные предыдущего процесса
    context.user_data.clear()
    await query.message.reply_text(
        "Для покупки терминала мне нужна ваша информация.\n\n"
        "Пожалуйста, введите ваш номер телефона (формат: 8XXXXXXXXXX или +7XXXXXXXXXX):"
    )
    return WAITING_PHONE_PURCHASE


async def button_callback_activate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    # Очищаем данные предыдущего процесса
    context.user_data.clear()
    await query.message.reply_text(
        "Для активации терминала мне нужна ваша информация.\n\n"
        "Пожалуйста, введите ваш номер телефона (формат: 8XXXXXXXXXX или +7XXXXXXXXXX):"
    )
    return WAITING_PHONE_ACTIVATE


async def handle_phone_purchase(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = normalize_phone(update.message.text)
    if not phone:
        await update.message.reply_text(
            "❌ Неверный формат номера телефона.\n\n"
            "Номер должен содержать 11 цифр и начинаться с +7 или 8.\n"
            "Пример: +79991234567 или 89991234567\n\n"
            "Пожалуйста, введите номер еще раз:"
        )
        return WAITING_PHONE_PURCHASE
    
    context.user_data['phone'] = phone
    await update.message.reply_text("Теперь введите ваше имя:")
    return WAITING_NAME_PURCHASE


async def handle_name_purchase(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    
    if not is_cyrillic_only(name):
        await update.message.reply_text(
            "❌ Имя должно содержать только русские буквы (кириллицу) и пробелы.\n\n"
            "Пожалуйста, введите ваше имя еще раз:"
        )
        return WAITING_NAME_PURCHASE
    
    user_id = update.effective_user.id
    phone = context.user_data['phone']
    username = update.effective_user.username  # Получаем username, если доступен
    
    purchase_id = add_purchase(user_id, phone, name, username)
    request_number = f"BUY-{purchase_id:06d}"  # Номер заявки в формате BUY-000001
    
    await update.message.reply_text(
        f"✅ Заявка создана!\n\n"
        f"Номер вашей заявки: <b>{request_number}</b>\n\n"
        f"Спасибо! Мы с вами свяжемся.",
        parse_mode='HTML'
    )
    context.user_data.clear()
    return ConversationHandler.END


async def handle_phone_activate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = normalize_phone(update.message.text)
    if not phone:
        await update.message.reply_text(
            "❌ Неверный формат номера телефона.\n\n"
            "Номер должен содержать 11 цифр и начинаться с +7 или 8.\n"
            "Пример: +79991234567 или 89991234567\n\n"
            "Пожалуйста, введите номер еще раз:"
        )
        return WAITING_PHONE_ACTIVATE
    
    context.user_data['phone'] = phone
    await update.message.reply_text("Теперь введите ваше имя:")
    return WAITING_NAME_ACTIVATE


async def handle_name_activate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    
    if not is_cyrillic_only(name):
        await update.message.reply_text(
            "❌ Имя должно содержать только русские буквы (кириллицу) и пробелы.\n\n"
            "Пожалуйста, введите ваше имя еще раз:"
        )
        return WAITING_NAME_ACTIVATE
    
    user_id = update.effective_user.id
    phone = context.user_data['phone']
    username = update.effective_user.username  # Получаем username, если доступен
    
    activation_id = add_activation(user_id, phone, name, username)
    request_number = f"ST-{activation_id:06d}"  # Номер заявки в формате ST-000001
    context.user_data['activation_id'] = activation_id
    context.user_data['name'] = name
    context.user_data['phone'] = phone
    context.user_data['request_number'] = request_number
    
    # Отправляем номер заявки пользователю
    await update.message.reply_text(
        f"✅ Заявка создана!\n\n"
        f"Номер вашей заявки: <b>{request_number}</b>\n\n"
        f"Сохраните этот номер для отслеживания статуса.",
        parse_mode='HTML'
    )
    
    message_text = (
        "Спасибо за доверие! Для активации от Вас нужен серийный номер "
        "(написан на ножке после букв SN) + фото серийного номера "
        "(чтобы исключить риск активации чужого устройства), прилагаем пример:"
    )
    
    photo_path_jpg = os.path.join(os.path.dirname(__file__), "images", "serial_number_example.jpg")
    photo_path_png = os.path.join(os.path.dirname(__file__), "images", "serial_number_example.png")
    
    photo_sent = False
    if os.path.exists(photo_path_jpg):
        try:
            with open(photo_path_jpg, 'rb') as photo:
                await update.message.reply_photo(
                    photo=photo,
                    caption=message_text
                )
            photo_sent = True
        except Exception as e:
            print(f"Ошибка отправки фото JPG: {e}")
    
    if not photo_sent and os.path.exists(photo_path_png):
        try:
            with open(photo_path_png, 'rb') as photo:
                await update.message.reply_photo(
                    photo=photo,
                    caption=message_text
                )
            photo_sent = True
        except Exception as e:
            print(f"Ошибка отправки фото PNG: {e}")
    
    if not photo_sent:
        await update.message.reply_text(message_text)
    
    await update.message.reply_text(
        "Пожалуйста, введите серийный номер устройства (SN):"
    )
    return WAITING_SERIAL


def is_cyrillic_only(text):
    """Проверяет, что текст содержит только кириллицу и пробелы."""
    cyrillic_letters = 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯабвгдеёжзийклмнопрстуфхцчшщъыьэюя '
    return all(char in cyrillic_letters for char in text) and len(text.strip()) > 0


def is_valid_serial_number(text):
    """Проверяет, что серийный номер содержит только латиницу и цифры или только цифры."""
    text = text.strip()
    if not text:
        return False
    # Проверяем: только латиница (A-Z, a-z) и цифры (0-9) ИЛИ только цифры
    return text.isalnum() and all(ord(char) < 128 for char in text)  # Только ASCII символы (латиница + цифры)


async def handle_serial_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    serial_number = update.message.text.strip()
    
    if not is_valid_serial_number(serial_number):
        await update.message.reply_text(
            "❌ Неверный формат серийного номера.\n\n"
            "Серийный номер должен содержать только латинские буквы и цифры, или только цифры.\n\n"
            "Пожалуйста, введите серийный номер еще раз:"
        )
        return WAITING_SERIAL
    
    user_id = update.effective_user.id
    update_activation_serial_number(user_id, serial_number)
    
    keyboard = [
        [InlineKeyboardButton("⏭️ Пропустить фото", callback_data="skip_serial_photo")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        "Теперь отправьте фото серийного номера:",
        reply_markup=reply_markup
    )
    return WAITING_SERIAL_PHOTO


async def handle_serial_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    file_id = None
    
    if update.message.photo:
        file_id = update.message.photo[-1].file_id
    elif update.message.document:
        file_id = update.message.document.file_id
    else:
        await update.message.reply_text(
            "Пожалуйста, отправьте фото серийного номера (фото или документ)."
        )
        return WAITING_SERIAL_PHOTO
    
    update_activation_serial_photo(user_id, file_id)
    
    message_text = (
        "А также серийный номер с коробки терминала (написан после букв SN) + его фото, "
        "прилагаем пример:"
    )
    
    photo_path_jpg = os.path.join(os.path.dirname(__file__), "images", "serial_number_box_example.jpg")
    photo_path_png = os.path.join(os.path.dirname(__file__), "images", "serial_number_box_example.png")
    
    photo_sent = False
    if os.path.exists(photo_path_jpg):
        try:
            with open(photo_path_jpg, 'rb') as photo:
                await update.message.reply_photo(
                    photo=photo,
                    caption=message_text
                )
            photo_sent = True
        except Exception as e:
            print(f"Ошибка отправки фото JPG: {e}")
    
    if not photo_sent and os.path.exists(photo_path_png):
        try:
            with open(photo_path_png, 'rb') as photo:
                await update.message.reply_photo(
                    photo=photo,
                    caption=message_text
                )
            photo_sent = True
        except Exception as e:
            print(f"Ошибка отправки фото PNG: {e}")
    
    if not photo_sent:
        await update.message.reply_text(message_text)
    
    await update.message.reply_text(
        "Пожалуйста, введите серийный номер с коробки (SN):"
    )
    return WAITING_BOX_SERIAL


async def skip_serial_photo_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик пропуска фото серийного номера устройства"""
    query = update.callback_query
    await query.answer()
    
    warning_text = (
        "⚠️ <b>Внимание!</b>\n\n"
        "Вы пропустили отправку фото серийного номера устройства.\n\n"
        "Если серийный номер будет указан неверно, вся ответственность за это ложится на вас.\n\n"
        "Продолжаем без фото..."
    )
    
    await query.message.reply_text(warning_text, parse_mode='HTML')
    
    # Продолжаем процесс - запрашиваем серийный номер с коробки
    message_text = (
        "А также серийный номер с коробки терминала (написан после букв SN) + его фото, "
        "прилагаем пример:"
    )
    
    photo_path_jpg = os.path.join(os.path.dirname(__file__), "images", "serial_number_box_example.jpg")
    photo_path_png = os.path.join(os.path.dirname(__file__), "images", "serial_number_box_example.png")
    
    photo_sent = False
    if os.path.exists(photo_path_jpg):
        try:
            with open(photo_path_jpg, 'rb') as photo:
                await query.message.reply_photo(
                    photo=photo,
                    caption=message_text
                )
            photo_sent = True
        except Exception as e:
            print(f"Ошибка отправки фото JPG: {e}")
    
    if not photo_sent and os.path.exists(photo_path_png):
        try:
            with open(photo_path_png, 'rb') as photo:
                await query.message.reply_photo(
                    photo=photo,
                    caption=message_text
                )
            photo_sent = True
        except Exception as e:
            print(f"Ошибка отправки фото PNG: {e}")
    
    if not photo_sent:
        await query.message.reply_text(message_text)
    
    await query.message.reply_text(
        "Пожалуйста, введите серийный номер с коробки (SN):"
    )
    return WAITING_BOX_SERIAL


async def handle_serial_photo_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Пожалуйста, отправьте фото серийного номера (фото или документ). "
        "Вы также можете отменить операцию командой /cancel"
    )
    return WAITING_SERIAL_PHOTO


async def handle_box_serial_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    box_serial_number = update.message.text.strip()
    
    if not is_valid_serial_number(box_serial_number):
        await update.message.reply_text(
            "❌ Неверный формат серийного номера.\n\n"
            "Серийный номер должен содержать только латинские буквы и цифры, или только цифры.\n\n"
            "Пожалуйста, введите серийный номер с коробки еще раз:"
        )
        return WAITING_BOX_SERIAL
    
    user_id = update.effective_user.id
    update_activation_box_serial_number(user_id, box_serial_number)
    
    keyboard = [
        [InlineKeyboardButton("⏭️ Пропустить фото", callback_data="skip_box_photo")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        "Теперь отправьте фото серийного номера с коробки:",
        reply_markup=reply_markup
    )
    return WAITING_BOX_SERIAL_PHOTO


async def skip_box_photo_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик пропуска фото серийного номера коробки"""
    query = update.callback_query
    await query.answer()
    
    warning_text = (
        "⚠️ <b>Внимание!</b>\n\n"
        "Вы пропустили отправку фото серийного номера с коробки.\n\n"
        "Если серийные номера будут указаны неверно, вся ответственность за это ложится на вас.\n\n"
        "✅ Все данные получены!\n\n"
        "Пожалуйста, ожидайте. ⏳\n\n"
        "Мы свяжемся с вами в ближайшее время."
    )
    
    await query.message.reply_text(warning_text, parse_mode='HTML')
    context.user_data.clear()
    return ConversationHandler.END


async def handle_box_serial_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    file_id = None
    
    if update.message.photo:
        file_id = update.message.photo[-1].file_id
    elif update.message.document:
        file_id = update.message.document.file_id
    else:
        await update.message.reply_text(
            "Пожалуйста, отправьте фото серийного номера с коробки (фото или документ)."
        )
        return WAITING_BOX_SERIAL_PHOTO
    
    update_activation_box_serial_photo(user_id, file_id)
    
    # После получения фото коробки завершаем и просим ожидать
    await update.message.reply_text(
        "✅ Все данные получены!\n\n"
        "Пожалуйста, ожидайте. ⏳\n\n"
        "Мы свяжемся с вами в ближайшее время."
    )
    context.user_data.clear()
    return ConversationHandler.END


async def handle_box_serial_photo_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Пожалуйста, отправьте фото серийного номера с коробки (фото или документ). "
        "Вы также можете отменить операцию командой /cancel"
    )
    return WAITING_BOX_SERIAL_PHOTO


async def handle_kit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kit_number = update.message.text.strip()
    user_id = update.effective_user.id
    
    update_activation_kit(user_id, kit_number)
    
    await update.message.reply_text(
        "KIT номер сохранен. Пожалуйста, ожидайте. ⏳"
    )
    context.user_data.clear()
    return ConversationHandler.END


async def precheckout_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.pre_checkout_query
    await query.answer(ok=True)
    
async def successful_payment_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    payment = update.message.successful_payment
    user_id = update.effective_user.id
    
    update_activation_receipt(user_id, payment.telegram_payment_charge_id)
    
    await update.message.reply_text(
        "✅ Платеж успешно получен!\n\n"
        "Пожалуйста, ожидайте. ⏳\n\n"
        "Мы свяжемся с вами в ближайшее время."
    )
    context.user_data.clear()
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Операция отменена.")
    context.user_data.clear()
    return ConversationHandler.END

async def start_fallback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    welcome_text = (
        "Добро пожаловать! 👋\n\n"
        "Это техподдержка по активации терминалов Starlink. "
        "Я помогу вам купить терминал или активировать уже имеющееся устройство.\n\n"
        "Выберите нужное действие:"
    )
    
    keyboard = [
        [InlineKeyboardButton("🛒 Купить терминал", callback_data="buy")],
        [InlineKeyboardButton("⚙️ Активировать", callback_data="activate")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(welcome_text, reply_markup=reply_markup)
    return ConversationHandler.END


def is_admin(user_id):
    return user_id in ADMIN_IDS


async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        return
    
    context.user_data['admin_auth'] = True
    await update.message.reply_text(
        "🔐 Админ-панель\n\nВведите пароль для доступа:"
    )
    return WAITING_ADMIN_PASSWORD


async def admin_password_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    if not is_admin(user_id) or not context.user_data.get('admin_auth'):
        return ConversationHandler.END
    
    password = update.message.text.strip()
    
    if password != ADMIN_PASSWORD:
        await update.message.reply_text("❌ Неверный пароль. Попробуйте еще раз:")
        return WAITING_ADMIN_PASSWORD
    
    context.user_data.pop('admin_auth', None)
    
    keyboard = [
        [InlineKeyboardButton("🔍 Поиск заявки", callback_data="admin_search")],
        [InlineKeyboardButton("📊 Статистика", callback_data="admin_stats")],
        [InlineKeyboardButton("🛒 Покупки", callback_data="admin_purchases")],
        [InlineKeyboardButton("⚙️ Активации", callback_data="admin_activations")],
        [InlineKeyboardButton("📋 Активации (детально)", callback_data="admin_activations_detail")],
        [InlineKeyboardButton("📄 Экспорт в Excel", callback_data="admin_export_excel")],
        [InlineKeyboardButton("✅ Отметить как обработанную", callback_data="admin_mark_processed")],
        [InlineKeyboardButton("✉️ Привязать Email/Пароль", callback_data="admin_add_credentials")],
        [InlineKeyboardButton("🚪 Выход из админ-панели", callback_data="admin_exit")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(
        "🔐 Админ-панель\n\nВыберите действие:",
        reply_markup=reply_markup
    )
    return ConversationHandler.END


async def admin_email_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        return ConversationHandler.END
    
    # Проверяем, что это действительно состояние для ввода email
    if context.user_data.get('admin_cred_state') != WAITING_ADMIN_EMAIL:
        return ConversationHandler.END
    
    email = update.message.text.strip()
    context.user_data['cred_email'] = email
    activation_id = context.user_data.get('cred_activation_id')
    
    if not activation_id:
        await update.message.reply_text("❌ Ошибка: не выбрана заявка.")
        context.user_data.pop('admin_cred_state', None)
        return ConversationHandler.END
    
    context.user_data['admin_cred_state'] = WAITING_ADMIN_PASSWORD_FIELD
    await update.message.reply_text(
        f"📝 Теперь введите пароль для заявки ST-{activation_id:06d}:"
    )
    return WAITING_ADMIN_PASSWORD_FIELD


async def admin_password_field_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        return ConversationHandler.END
    
    # Проверяем, что это действительно состояние для ввода пароля
    if context.user_data.get('admin_cred_state') != WAITING_ADMIN_PASSWORD_FIELD:
        return ConversationHandler.END
    
    password = update.message.text.strip()
    activation_id = context.user_data.get('cred_activation_id')
    email = context.user_data.get('cred_email')
    
    if not activation_id or not email:
        await update.message.reply_text("❌ Ошибка: не заполнены все данные.")
        context.user_data.pop('cred_activation_id', None)
        context.user_data.pop('cred_email', None)
        context.user_data.pop('admin_cred_state', None)
        return ConversationHandler.END
    
    if update_activation_email_password(activation_id, email, password):
        request_number = f"ST-{activation_id:06d}"
        await update.message.reply_text(
            f"✅ Email и пароль успешно привязаны к заявке {request_number}!"
        )
        
        # НЕ отправляем пользователю email и пароль - это только для админа
        
        context.user_data.pop('cred_activation_id', None)
        context.user_data.pop('cred_email', None)
        context.user_data.pop('admin_cred_state', None)
        
        # Показываем обновленную заявку
        activation = get_activation_by_id(activation_id)
        if activation:
            await show_activation_details(update, context, activation)
        
        return ConversationHandler.END
    else:
        await update.message.reply_text(f"❌ Ошибка при сохранении данных.")
        context.user_data.pop('cred_activation_id', None)
        context.user_data.pop('cred_email', None)
        context.user_data.pop('admin_cred_state', None)
        return ConversationHandler.END


async def show_activation_details(update: Update, context: ContextTypes.DEFAULT_TYPE, activation):
    """Универсальная функция для показа детальной информации о заявке активации с кнопками редактирования"""
    act_id, uid, phone, name, username, created_at, payment, receipt, serial_num, serial_photo, box_serial, box_photo, kit, status, service_provided, service_provided_at, email, password = activation[:18]
    request_number = f"ST-{act_id:06d}"
    
    text = f"📋 Детальная информация по заявке {request_number}\n\n"
    text += f"🔹 ID заявки: {act_id}\n"
    text += f"User ID: {uid}\n"
    text += f"Username: @{username}\n" if username else "Username: не указан\n"
    text += f"Имя: {name}\n"
    text += f"Телефон: {phone}\n"
    text += f"Дата создания: {created_at[:19]}\n"
    text += f"Статус: {status}\n"
    text += f"Оплата получена: {'✅ Да' if payment else '❌ Нет'}\n"
    text += f"Услуга оказана: {'✅ Да' if service_provided else '❌ Нет'}\n"
    
    if service_provided_at:
        start_date = datetime.fromisoformat(service_provided_at)
        end_date = start_date + timedelta(days=30)
        text += f"Дата начала активации: {service_provided_at[:19]}\n"
        text += f"Дата окончания подписки: {end_date.strftime('%Y-%m-%d %H:%M:%S')}\n"
    
    text += f"\n📦 Данные устройства:\n"
    text += f"SN устройство: {serial_num if serial_num else 'не указан'}\n"
    text += f"SN коробка: {box_serial if box_serial else 'не указан'}\n"
    if kit:
        text += f"KIT номер: {kit}\n"
    
    if email:
        text += f"\n📧 Email: {email}\n"
    if password:
        text += f"🔑 Пароль: {password}\n"
    
    # Кнопки редактирования
    keyboard = []
    keyboard.append([InlineKeyboardButton("✏️ Редактировать Email/Пароль", callback_data=f"edit_cred_{act_id}")])
    
    if service_provided:
        keyboard.append([InlineKeyboardButton("❌ Снять отметку об обработке", callback_data=f"toggle_status_{act_id}")])
    else:
        keyboard.append([InlineKeyboardButton("✅ Отметить как обработанную", callback_data=f"toggle_status_{act_id}")])
    
    keyboard.append([InlineKeyboardButton("🗑️ Удалить заявку", callback_data=f"delete_confirm_{act_id}")])
    
    # Определяем, откуда пришли (поиск или список)
    back_to = context.user_data.get('admin_view_back_to', 'admin_activations')
    keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data=back_to)])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    # Если это query (callback), используем редактирование сообщения, иначе отправляем новое
    if hasattr(update, 'callback_query') and update.callback_query:
        await update.callback_query.message.reply_text(text, reply_markup=reply_markup)
    else:
        await update.message.reply_text(text, reply_markup=reply_markup)
    
    # Генерируем и отправляем Excel файл с данными заявки
    if hasattr(update, 'callback_query') and update.callback_query:
        await update.callback_query.message.reply_text("📄 Генерирую Excel файл...")
    else:
        await update.message.reply_text("📄 Генерирую Excel файл...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Активация"
    
    headers = ["Номер заявки", "User ID", "Username", "Номер телефона", "Имя", "Дата заявки", "Услуга",
               "SN устройство", "SN коробка", "KIT номер",
               "Дата начала активации", "Дата окончания подписки", "Email", "Пароль"]
    ws.append(headers)
    
    # Форматирование заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Добавляем данные заявки
    start_date_str = ""
    end_date_str = ""
    
    if service_provided_at:
        start_date = datetime.fromisoformat(service_provided_at)
        end_date = start_date + timedelta(days=30)
        start_date_str = start_date.strftime('%Y-%m-%d %H:%M:%S')
        end_date_str = end_date.strftime('%Y-%m-%d %H:%M:%S')
    
    ws.append([
        request_number,
        uid,
        f"@{username}" if username else "",
        phone,
        name,
        created_at[:19],
        "Активация",
        serial_num if serial_num else "",
        box_serial if box_serial else "",
        kit if kit else "",
        start_date_str,
        end_date_str,
        email if email else "",
        password if password else ""
    ])
    
    # Автоматическая ширина столбцов
    from openpyxl.utils import get_column_letter
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value:
                cell_value = str(cell.value)
                max_length = max(max_length, len(cell_value))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    filename = f"activation_{request_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    if hasattr(update, 'callback_query') and update.callback_query:
        await update.callback_query.message.reply_document(
            document=open(filename, 'rb'),
            filename=filename
        )
    else:
        await update.message.reply_document(
            document=open(filename, 'rb'),
            filename=filename
        )
    
    os.remove(filename)


async def show_activation_details(update: Update, context: ContextTypes.DEFAULT_TYPE, activation):
    """Универсальная функция для показа детальной информации о заявке активации с кнопками редактирования"""
    act_id, uid, phone, name, username, created_at, payment, receipt, serial_num, serial_photo, box_serial, box_photo, kit, status, service_provided, service_provided_at, email, password = activation[:18]
    request_number = f"ST-{act_id:06d}"
    
    text = f"📋 Детальная информация по заявке {request_number}\n\n"
    text += f"🔹 ID заявки: {act_id}\n"
    text += f"User ID: {uid}\n"
    username_str = f"@{username}" if username else "не указан"
    text += f"Username: {username_str}\n"
    text += f"Имя: {name}\n"
    text += f"Телефон: {phone}\n"
    text += f"Дата создания: {created_at[:19]}\n"
    text += f"Статус: {status}\n"
    text += f"Оплата получена: {'✅ Да' if payment else '❌ Нет'}\n"
    text += f"Услуга оказана: {'✅ Да' if service_provided else '❌ Нет'}\n"
    
    if service_provided_at:
        start_date = datetime.fromisoformat(service_provided_at)
        end_date = start_date + timedelta(days=30)
        text += f"Дата начала активации: {service_provided_at[:19]}\n"
        text += f"Дата окончания подписки: {end_date.strftime('%Y-%m-%d %H:%M:%S')}\n"
    
    text += f"\n📦 Данные устройства:\n"
    text += f"SN устройство: {serial_num if serial_num else 'не указан'}\n"
    text += f"SN коробка: {box_serial if box_serial else 'не указан'}\n"
    if kit:
        text += f"KIT номер: {kit}\n"
    
    if email:
        text += f"\n📧 Email: {email}\n"
    if password:
        text += f"🔑 Пароль: {password}\n"
    
    # Кнопки редактирования
    keyboard = []
    keyboard.append([InlineKeyboardButton("✏️ Редактировать Email/Пароль", callback_data=f"edit_cred_{act_id}")])
    
    if service_provided:
        keyboard.append([InlineKeyboardButton("❌ Снять отметку об обработке", callback_data=f"toggle_status_{act_id}")])
    else:
        keyboard.append([InlineKeyboardButton("✅ Отметить как обработанную", callback_data=f"toggle_status_{act_id}")])
    
    keyboard.append([InlineKeyboardButton("🗑️ Удалить заявку", callback_data=f"delete_confirm_{act_id}")])
    
    # Определяем, откуда пришли (поиск или список)
    back_to = context.user_data.get('admin_view_back_to', 'admin_activations')
    keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data=back_to)])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    # Если это query (callback), используем редактирование сообщения, иначе отправляем новое
    if hasattr(update, 'callback_query') and update.callback_query:
        query = update.callback_query
        await query.message.reply_text(text, reply_markup=reply_markup)
        msg_for_excel = query.message
    else:
        await update.message.reply_text(text, reply_markup=reply_markup)
        msg_for_excel = update.message
    
    # Генерируем и отправляем Excel файл с данными заявки
    await msg_for_excel.reply_text("📄 Генерирую Excel файл...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Активация"
    
    headers = ["Номер заявки", "User ID", "Username", "Номер телефона", "Имя", "Дата заявки", "Услуга",
               "SN устройство", "SN коробка", "KIT номер",
               "Дата начала активации", "Дата окончания подписки", "Email", "Пароль"]
    ws.append(headers)
    
    # Форматирование заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Добавляем данные заявки
    start_date_str = ""
    end_date_str = ""
    
    if service_provided_at:
        start_date = datetime.fromisoformat(service_provided_at)
        end_date = start_date + timedelta(days=30)
        start_date_str = start_date.strftime('%Y-%m-%d %H:%M:%S')
        end_date_str = end_date.strftime('%Y-%m-%d %H:%M:%S')
    
    ws.append([
        request_number,
        uid,
        f"@{username}" if username else "",
        phone,
        name,
        created_at[:19],
        "Активация",
        serial_num if serial_num else "",
        box_serial if box_serial else "",
        kit if kit else "",
        start_date_str,
        end_date_str,
        email if email else "",
        password if password else ""
    ])
    
    # Автоматическая ширина столбцов
    from openpyxl.utils import get_column_letter
    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value:
                cell_value = str(cell.value)
                max_length = max(max_length, len(cell_value))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    filename = f"activation_{request_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    await msg_for_excel.reply_document(
        document=open(filename, 'rb'),
        filename=filename
    )
    
    os.remove(filename)


async def admin_search_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик поиска заявки по номеру"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        return ConversationHandler.END
    
    request_number = update.message.text.strip().upper()
    
    # Ищем активацию
    activation = find_activation_by_request_number(request_number)
    if activation:
        context.user_data['admin_view_back_to'] = 'admin_search_back'
        await show_activation_details(update, context, activation)
        return ConversationHandler.END
    
    # Ищем покупку
    purchase = find_purchase_by_request_number(request_number)
    if purchase:
        pur_id, uid, phone, name, username, created_at = purchase
        request_number_formatted = f"BUY-{pur_id:06d}"
        text = f"📋 Детальная информация по заявке {request_number_formatted}\n\n"
        text += f"🔹 ID заявки: {pur_id}\n"
        text += f"User ID: {uid}\n"
        username_str = f"@{username}" if username else "не указан"
        text += f"Username: {username_str}\n"
        text += f"Имя: {name}\n"
        text += f"Телефон: {phone}\n"
        text += f"Дата создания: {created_at[:19]}\n"
        
        keyboard = [
            [InlineKeyboardButton("🗑️ Удалить заявку", callback_data=f"delete_purchase_{pur_id}")],
            [InlineKeyboardButton("🔙 Назад", callback_data="admin_search_back")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(text, reply_markup=reply_markup)
        return ConversationHandler.END
    
    # Не найдено
    await update.message.reply_text(
        f"❌ Заявка с номером {request_number} не найдена.\n\n"
        f"Попробуйте еще раз или отправьте /cancel для отмены."
    )
    return WAITING_ADMIN_SEARCH


async def admin_search_callback_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Entry point для поиска из CallbackQuery"""
    if update.callback_query and update.callback_query.data == "admin_search":
        user_id = update.effective_user.id
        if is_admin(user_id):
            await update.callback_query.answer()
            await update.callback_query.message.reply_text(
                "🔍 Введите номер заявки для поиска:\n\n"
                "Формат: ST-000001 (для активаций) или BUY-000001 (для покупок)\n\n"
                "Или отправьте /cancel для отмены."
            )
            return WAITING_ADMIN_SEARCH
    return None


async def admin_edit_callback_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Entry point для редактирования email/пароля из CallbackQuery"""
    if update.callback_query and update.callback_query.data and update.callback_query.data.startswith("edit_cred_"):
        user_id = update.effective_user.id
        if is_admin(user_id):
            await update.callback_query.answer()
            activation_id = int(update.callback_query.data.split("_")[2])
            context.user_data['cred_activation_id'] = activation_id
            context.user_data['admin_cred_state'] = WAITING_ADMIN_EMAIL
            activation = get_activation_by_id(activation_id)
            if activation:
                act_id, uid, phone, name, username, created_at, payment, receipt, serial_num, serial_photo, box_serial, box_photo, kit, status, service_provided, service_provided_at, email, password = activation[:18]
                request_number = f"ST-{act_id:06d}"
                current_info = f"\nТекущий email: {email if email else 'не указан'}\nТекущий пароль: {'*' * len(password) if password else 'не указан'}" if email or password else ""
                await update.callback_query.message.reply_text(
                    f"📝 Введите email для заявки {request_number} ({name}):{current_info}\n\n"
                    f"Или отправьте /cancel для отмены."
                )
                return WAITING_ADMIN_EMAIL
    return None


async def admin_start_fallback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Fallback для /start в админ ConversationHandler"""
    context.user_data.clear()
    welcome_text = (
        "Добро пожаловать! 👋\n\n"
        "Это техподдержка по активации терминалов Starlink. "
        "Я помогу вам купить терминал или активировать уже имеющееся устройство.\n\n"
        "Выберите нужное действие:"
    )
    
    keyboard = [
        [InlineKeyboardButton("🛒 Купить терминал", callback_data="buy")],
        [InlineKeyboardButton("⚙️ Активировать", callback_data="activate")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await update.message.reply_text(welcome_text, reply_markup=reply_markup)
    return ConversationHandler.END


async def admin_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = update.effective_user.id
    if not is_admin(user_id):
        await query.message.reply_text("❌ У вас нет доступа.")
        return
    
    if query.data == "admin_stats":
        stats = get_statistics()
        text = (
            f"📊 Статистика\n\n"
            f"🛒 Всего покупок: {stats['total_purchases']}\n"
            f"⚙️ Всего активаций: {stats['total_activations']}\n\n"
            f"⏳ Ожидают оплаты: {stats['pending_activations']}\n"
            f"💳 Оплата подтверждена: {stats['payment_confirmed']}\n"
            f"✅ Завершено: {stats['completed_activations']}"
        )
        await query.message.reply_text(text)
    
    elif query.data == "admin_purchases":
        purchases = get_all_purchases()
        if not purchases:
            await query.message.reply_text("📭 Покупок пока нет.")
            return
        
        text = "🛒 Все покупки:\n\n"
        for purchase in purchases[:20]:
            purchase_id, uid, phone, name, username, created_at = purchase[:6]
            username_str = f"@{username}" if username else "не указан"
            text += (
                f"ID: {purchase_id}\n"
                f"User ID: {uid}\n"
                f"Username: {username_str}\n"
                f"Имя: {name}\n"
                f"Телефон: {phone}\n"
                f"Дата: {created_at[:19]}\n"
                f"{'─' * 30}\n"
            )
        
        if len(purchases) > 20:
            text += f"\n... и еще {len(purchases) - 20} записей"
        
        await query.message.reply_text(text)
    
    elif query.data == "admin_activations":
        # Показываем две кнопки: Ожидают и Обработанные
        keyboard = [
            [InlineKeyboardButton("⏳ Ожидают", callback_data="admin_activations_pending_page_0")],
            [InlineKeyboardButton("✅ Обработанные", callback_data="admin_activations_processed_page_0")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.reply_text(
            "⚙️ Выберите категорию активаций:",
            reply_markup=reply_markup
        )
    
    elif query.data == "admin_activations_detail":
        activations = get_all_activations()
        if not activations:
            await query.message.reply_text("📭 Активаций пока нет.")
            return
        
        text = "📋 Детальная информация по активациям:\n\n"
        for act in activations[:10]:
            act_id, uid, phone, name, username, created_at, payment, receipt, serial_num, serial_photo, box_serial, box_photo, kit, status, service_provided, service_provided_at = act[:16]
            username_str = f"@{username}" if username else "не указан"
            text += (
                f"🔹 ID заявки: {act_id}\n"
                f"User ID: {uid}\n"
                f"Username: {username_str}\n"
                f"Имя: {name}\n"
                f"Телефон: {phone}\n"
                f"Статус: {status}\n"
                f"Оплата получена: {'Да' if payment else 'Нет'}\n"
                f"SN устройство: {serial_num if serial_num else 'не указан'}\n"
                f"SN коробка: {box_serial if box_serial else 'не указан'}\n"
                f"KIT номер: {kit if kit else 'не указан'}\n"
                f"Услуга оказана: {'✅ Да' if service_provided else '❌ Нет'}\n"
                f"Дата создания: {created_at[:19]}\n"
            )
            if service_provided_at:
                start_date = datetime.fromisoformat(service_provided_at)
                end_date = start_date + timedelta(days=30)
                text += f"Дата начала активации: {service_provided_at[:19]}\n"
                text += f"Дата окончания подписки: {end_date.strftime('%Y-%m-%d %H:%M:%S')}\n"
            text += f"{'═' * 35}\n"
        
        if len(activations) > 10:
            text += f"\n... и еще {len(activations) - 10} записей"
        
        await query.message.reply_text(text)
    
    elif query.data == "admin_export_excel":
        await query.message.reply_text("📄 Генерирую Excel файл...")
        activations = get_all_activations()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Активации"
        
        headers = ["Номер заявки", "User ID", "Username", "Номер телефона", "Имя", "Дата заявки", "Услуга",
                   "SN устройство", "SN коробка", "KIT номер",
                   "Дата начала активации", "Дата окончания подписки", "Email", "Пароль"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for act in activations:
            act_id, uid, phone, name, username, created_at, payment, receipt, serial_num, serial_photo, box_serial, box_photo, kit, status, service_provided, service_provided_at, email, password = act[:18]
            
            request_number = f"ST-{act_id:06d}"
            start_date_str = ""
            end_date_str = ""
            
            if service_provided_at:
                start_date = datetime.fromisoformat(service_provided_at)
                end_date = start_date + timedelta(days=30)
                start_date_str = start_date.strftime('%Y-%m-%d %H:%M:%S')
                end_date_str = end_date.strftime('%Y-%m-%d %H:%M:%S')
            
            ws.append([
                request_number,
                uid,
                f"@{username}" if username else "",
                phone,
                name,
                created_at[:19],
                "Активация",
                serial_num if serial_num else "",
                box_serial if box_serial else "",
                kit if kit else "",
                start_date_str,
                end_date_str,
                email if email else "",
                password if password else ""
            ])
        
        # Автоматическая ширина столбцов
        from openpyxl.utils import get_column_letter
        for col_idx, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                if cell.value:
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))
            # Устанавливаем ширину: длина контента + небольшой отступ, но не более 50 символов
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        filename = f"activations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        await query.message.reply_document(
            document=open(filename, 'rb'),
            filename=filename
        )
        
        os.remove(filename)
    
    elif query.data == "admin_mark_processed":
        activations = get_all_activations()
        if not activations:
            await query.message.reply_text("📭 Активаций пока нет.")
            return
        
        buttons = []
        for act in activations[:50]:
            act_id, uid, phone, name, username, created_at, payment, receipt, serial_num, serial_photo, box_serial, box_photo, kit, status, service_provided, service_provided_at, email, password = act[:18]
            if not service_provided:
                request_number = f"ST-{act_id:06d}"
                buttons.append([InlineKeyboardButton(
                    f"{request_number}: {name} ({phone})",
                    callback_data=f"mark_{act_id}"
                )])
        
        if not buttons:
            await query.message.reply_text("✅ Все заявки уже обработаны.")
            return
        
        reply_markup = InlineKeyboardMarkup(buttons)
        await query.message.reply_text(
            "Выберите заявку для отметки как обработанную:",
            reply_markup=reply_markup
        )
    
    elif query.data.startswith("mark_"):
        activation_id = int(query.data.split("_")[1])
        if mark_service_provided(activation_id):
            request_number = f"ST-{activation_id:06d}"
            await query.message.reply_text(f"✅ Заявка {request_number} отмечена как обработанная.")
        else:
            await query.message.reply_text(f"❌ Ошибка при обработке заявки #{activation_id}.")
    
    elif query.data == "admin_add_credentials":
        activations = get_all_activations()
        if not activations:
            await query.message.reply_text("📭 Активаций пока нет.")
            return
        
        buttons = []
        for act in activations[:50]:
            act_id, uid, phone, name, username, created_at, payment, receipt, serial_num, serial_photo, box_serial, box_photo, kit, status, service_provided, service_provided_at, email, password = act[:18]
            request_number = f"ST-{act_id:06d}"
            buttons.append([InlineKeyboardButton(
                f"{request_number}: {name} ({phone})" + (" ✉️" if email else ""),
                callback_data=f"add_cred_{act_id}"
            )])
        
        reply_markup = InlineKeyboardMarkup(buttons)
        await query.message.reply_text(
            "Выберите заявку для привязки email и пароля:",
            reply_markup=reply_markup
        )
    
    elif query.data.startswith("add_cred_"):
        activation_id = int(query.data.split("_")[2])
        context.user_data['cred_activation_id'] = activation_id
        context.user_data['admin_cred_state'] = WAITING_ADMIN_EMAIL
        activation = get_activation_by_id(activation_id)
        if activation:
            act_id, uid, phone, name, username, created_at, payment, receipt, serial_num, serial_photo, box_serial, box_photo, kit, status, service_provided, service_provided_at, email, password = activation[:18]
            request_number = f"ST-{act_id:06d}"
            current_info = f"\nТекущий email: {email if email else 'не указан'}\nТекущий пароль: {'*' * len(password) if password else 'не указан'}" if email or password else ""
            await query.message.reply_text(
                f"📝 Введите email для заявки {request_number} ({name}):{current_info}\n\n"
                f"Или отправьте /cancel для отмены."
            )
    
    elif query.data.startswith("admin_activations_pending_page_"):
        # Показываем список ожидающих заявок с пагинацией
        page = int(query.data.split("_")[-1])
        activations = get_pending_activations()
        
        if not activations:
            await query.message.reply_text("📭 Ожидающих заявок пока нет.")
            return
        
        buttons = []
        start_idx = page * 10
        end_idx = start_idx + 10
        
        for act in activations[start_idx:end_idx]:
            act_id, uid, phone, name = act[0], act[1], act[2], act[3]
            request_number = f"ST-{act_id:06d}"
            buttons.append([InlineKeyboardButton(
                f"{request_number}: {name} ({phone})",
                callback_data=f"view_activation_{act_id}"
            )])
        
        # Кнопки пагинации
        nav_buttons = []
        if page > 0:
            nav_buttons.append(InlineKeyboardButton("◀️ Назад", callback_data=f"admin_activations_pending_page_{page-1}"))
        if end_idx < len(activations):
            nav_buttons.append(InlineKeyboardButton("▶️ Вперед", callback_data=f"admin_activations_pending_page_{page+1}"))
        
        if nav_buttons:
            buttons.append(nav_buttons)
        
        buttons.append([InlineKeyboardButton("🔙 Назад к категориям", callback_data="admin_activations")])
        
        reply_markup = InlineKeyboardMarkup(buttons)
        total = len(activations)
        text = f"⏳ Ожидающие заявки (страница {page + 1})\n\n"
        text += f"Всего: {total} заявок\n"
        text += f"Показано: {start_idx + 1}-{min(end_idx, total)} из {total}\n\n"
        text += "Выберите заявку для просмотра деталей:"
        
        await query.message.reply_text(text, reply_markup=reply_markup)
    
    elif query.data.startswith("admin_activations_processed_page_"):
        # Показываем список обработанных заявок с пагинацией
        page = int(query.data.split("_")[-1])
        activations = get_processed_activations()
        
        if not activations:
            await query.message.reply_text("📭 Обработанных заявок пока нет.")
            return
        
        buttons = []
        start_idx = page * 10
        end_idx = start_idx + 10
        
        for act in activations[start_idx:end_idx]:
            act_id, uid, phone, name = act[0], act[1], act[2], act[3]
            request_number = f"ST-{act_id:06d}"
            buttons.append([InlineKeyboardButton(
                f"{request_number}: {name} ({phone})",
                callback_data=f"view_activation_{act_id}"
            )])
        
        # Кнопки пагинации
        nav_buttons = []
        if page > 0:
            nav_buttons.append(InlineKeyboardButton("◀️ Назад", callback_data=f"admin_activations_processed_page_{page-1}"))
        if end_idx < len(activations):
            nav_buttons.append(InlineKeyboardButton("▶️ Вперед", callback_data=f"admin_activations_processed_page_{page+1}"))
        
        if nav_buttons:
            buttons.append(nav_buttons)
        
        buttons.append([InlineKeyboardButton("🔙 Назад к категориям", callback_data="admin_activations")])
        
        reply_markup = InlineKeyboardMarkup(buttons)
        total = len(activations)
        text = f"✅ Обработанные заявки (страница {page + 1})\n\n"
        text += f"Всего: {total} заявок\n"
        text += f"Показано: {start_idx + 1}-{min(end_idx, total)} из {total}\n\n"
        text += "Выберите заявку для просмотра деталей:"
        
        await query.message.reply_text(text, reply_markup=reply_markup)
    
    elif query.data.startswith("view_activation_"):
        # Показываем детальную информацию о заявке
        activation_id = int(query.data.split("_")[2])
        activation = get_activation_by_id(activation_id)
        
        if not activation:
            await query.message.reply_text("❌ Заявка не найдена.")
            return
        
        # Сохраняем откуда пришли для кнопки "Назад" (если не установлено)
        if 'admin_view_back_to' not in context.user_data:
            context.user_data['admin_view_back_to'] = "admin_activations"
        await show_activation_details(update, context, activation)
    
    elif query.data.startswith("edit_cred_"):
        # Это теперь обрабатывается через entry point ConversationHandler
        pass
    
    elif query.data.startswith("toggle_status_"):
        activation_id = int(query.data.split("_")[2])
        if toggle_service_provided(activation_id):
            # Обновляем вид заявки
            activation = get_activation_by_id(activation_id)
            if activation:
                status_text = "отмечена как обработанная" if activation[14] else "отметка снята"
                request_number = f"ST-{activation[0]:06d}"
                await query.message.reply_text(f"✅ Заявка {request_number} {status_text}.")
                # Показываем обновленную заявку
                context.user_data['admin_view_back_to'] = context.user_data.get('admin_view_back_to', 'admin_activations')
                await show_activation_details(update, context, activation)
        else:
            await query.message.reply_text(f"❌ Ошибка при изменении статуса заявки.")
    
    elif query.data.startswith("delete_confirm_"):
        activation_id = int(query.data.split("_")[2])
        activation = get_activation_by_id(activation_id)
        if activation:
            act_id = activation[0]
            request_number = f"ST-{act_id:06d}"
            keyboard = [
                [InlineKeyboardButton("✅ Да, удалить", callback_data=f"delete_yes_{act_id}")],
                [InlineKeyboardButton("❌ Отмена", callback_data=f"view_activation_{act_id}")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.message.reply_text(
                f"⚠️ Вы уверены, что хотите удалить заявку {request_number}?\n\n"
                f"Это действие нельзя отменить!",
                reply_markup=reply_markup
            )
    
    elif query.data.startswith("delete_yes_"):
        activation_id = int(query.data.split("_")[2])
        request_number = f"ST-{activation_id:06d}"
        if delete_activation(activation_id):
            await query.message.reply_text(f"✅ Заявка {request_number} успешно удалена.")
        else:
            await query.message.reply_text(f"❌ Ошибка при удалении заявки {request_number}.")
    
    elif query.data.startswith("delete_purchase_"):
        purchase_id = int(query.data.split("_")[2])
        request_number = f"BUY-{purchase_id:06d}"
        if delete_purchase(purchase_id):
            await query.message.reply_text(f"✅ Заявка {request_number} успешно удалена.")
        else:
            await query.message.reply_text(f"❌ Ошибка при удалении заявки {request_number}.")
    
    elif query.data == "admin_exit":
        welcome_text = (
            "👋 Вы вышли из админ-панели.\n\n"
            "Добро пожаловать! 👋\n\n"
            "Это техподдержка по активации терминалов Starlink. "
            "Я помогу вам купить терминал или активировать уже имеющееся устройство.\n\n"
            "Выберите нужное действие:"
        )
        
        keyboard = [
            [InlineKeyboardButton("🛒 Купить терминал", callback_data="buy")],
            [InlineKeyboardButton("⚙️ Активировать", callback_data="activate")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.message.reply_text(welcome_text, reply_markup=reply_markup)
        return


def main():
    import sys
    sys.stdout.write("MAIN: Функция main() вызвана\n")
    sys.stdout.flush()
    sys.stderr.write("MAIN: Функция main() вызвана (stderr)\n")
    sys.stderr.flush()
    
    print("Инициализация базы данных...")
    try:
        init_database()
        print("База данных инициализирована")
    except Exception as e:
        print(f"Ошибка при инициализации базы данных: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    print("Создание Application...")
    try:
        application = Application.builder().token(BOT_TOKEN).build()
        print("Application создан")
    except Exception as e:
        print(f"Ошибка при создании Application: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    async def end_purchase_and_start_activate(update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Завершает процесс покупки и начинает процесс активации"""
        context.user_data.clear()
        query = update.callback_query
        await query.answer()
        # Завершаем текущий процесс и запускаем обработчик активации вручную
        await button_callback_activate(update, context)
        return ConversationHandler.END
    
    purchase_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(button_callback_buy, pattern="^buy$")],
        states={
            WAITING_PHONE_PURCHASE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_phone_purchase)
            ],
            WAITING_NAME_PURCHASE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_name_purchase)
            ],
        },
        fallbacks=[
            CallbackQueryHandler(end_purchase_and_start_activate, pattern="^activate$"),
            CommandHandler("cancel", cancel),
            CommandHandler("start", start_fallback)
        ],
        allow_reentry=True,
    )
    
    async def end_activate_and_start_purchase(update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Завершает процесс активации и начинает процесс покупки"""
        context.user_data.clear()
        query = update.callback_query
        await query.answer()
        # Завершаем текущий процесс и запускаем обработчик покупки вручную
        await button_callback_buy(update, context)
        return ConversationHandler.END
    
    activation_handler = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(button_callback_activate, pattern="^activate$")
        ],
        states={
            WAITING_PHONE_ACTIVATE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_phone_activate)
            ],
            WAITING_NAME_ACTIVATE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_name_activate)
            ],
            WAITING_SERIAL: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_serial_number)
            ],
            WAITING_SERIAL_PHOTO: [
                CallbackQueryHandler(skip_serial_photo_callback, pattern="^skip_serial_photo$"),
                MessageHandler(filters.PHOTO | filters.Document.ALL, handle_serial_photo),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_serial_photo_text)
            ],
            WAITING_BOX_SERIAL: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_box_serial_number)
            ],
            WAITING_BOX_SERIAL_PHOTO: [
                CallbackQueryHandler(skip_box_photo_callback, pattern="^skip_box_photo$"),
                MessageHandler(filters.PHOTO | filters.Document.ALL, handle_box_serial_photo),
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_box_serial_photo_text)
            ],
        },
        fallbacks=[
            CallbackQueryHandler(end_activate_and_start_purchase, pattern="^buy$"),
            CommandHandler("cancel", cancel),
            CommandHandler("start", start_fallback)
        ],
        allow_reentry=True,
    )
    
    admin_password_handler_conv = ConversationHandler(
        entry_points=[
            CommandHandler("admin", admin_command),
            CallbackQueryHandler(admin_search_callback_entry, pattern="^admin_search$"),
            CallbackQueryHandler(admin_edit_callback_entry, pattern="^edit_cred_")
        ],
        states={
            WAITING_ADMIN_PASSWORD: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_password_handler)
            ],
            WAITING_ADMIN_EMAIL: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_email_handler)
            ],
            WAITING_ADMIN_PASSWORD_FIELD: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_password_field_handler)
            ],
            WAITING_ADMIN_SEARCH: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, admin_search_handler)
            ],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CommandHandler("start", admin_start_fallback)
        ],
    )
    
    async def check_subscriptions(context: ContextTypes.DEFAULT_TYPE):
        activations = get_activations_for_subscription_reminders()
        now = datetime.now()
        
        for act in activations:
            act_id, user_id, phone, name, service_provided_at, last_reminder_day = act
            
            if not service_provided_at:
                continue
            
            try:
                start_date = datetime.fromisoformat(service_provided_at)
                end_date = start_date + timedelta(days=30)
                days_left = (end_date - now).days
                
                if 1 <= days_left <= 5:
                    if last_reminder_day != days_left:
                        reminder_text = (
                            f"⏰ Напоминание о подписке\n\n"
                            f"Ваша подписка Starlink заканчивается через {days_left} день(дня/дней).\n"
                            f"Дата окончания: {end_date.strftime('%d.%m.%Y')}\n\n"
                            f"Пожалуйста, продлите подписку."
                        )
                        try:
                            await context.bot.send_message(chat_id=user_id, text=reminder_text)
                            update_last_reminder_day(act_id, days_left)
                        except Exception as e:
                            print(f"Ошибка отправки напоминания пользователю {user_id}: {e}")
            except Exception as e:
                print(f"Ошибка обработки активации {act_id}: {e}")
    
    try:
        print("Настройка job_queue...")
        job_queue = application.job_queue
        if job_queue:
            job_queue.run_repeating(check_subscriptions, interval=3600, first=10)
        print("job_queue настроен")
        
        print("Регистрация обработчиков...")
        # Группа -1 для команд (высший приоритет)
        application.add_handler(CommandHandler("start", start), group=-1)
        print("Обработчик /start зарегистрирован")
        
        # Группа 0 для остальных обработчиков
        application.add_handler(PreCheckoutQueryHandler(precheckout_callback))
        application.add_handler(MessageHandler(filters.SUCCESSFUL_PAYMENT, successful_payment_callback))
        application.add_handler(CallbackQueryHandler(admin_callback, pattern="^(admin_(?!search)|mark_|add_cred_|view_activation_|toggle_status_|delete_confirm_|delete_yes_|delete_purchase_|admin_search_back)"))
        application.add_handler(admin_password_handler_conv)
        application.add_handler(purchase_handler)
        application.add_handler(activation_handler)
        print("Все обработчики зарегистрированы")
        
        print("Бот запущен...")
        application.run_polling(allowed_updates=Update.ALL_TYPES)
    except Exception as e:
        print(f"Ошибка при настройке обработчиков: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    import sys
    sys.stdout.write("START: Скрипт запущен\n")
    sys.stdout.flush()
    sys.stderr.write("START: Скрипт запущен (stderr)\n")
    sys.stderr.flush()
    
    print("START: __name__ == '__main__'")
    
    try:
        main()
    except Exception as e:
        error_msg = f"КРИТИЧЕСКАЯ ОШИБКА при запуске бота: {e}"
        print(error_msg)
        sys.stderr.write(error_msg + "\n")
        sys.stderr.flush()
        import traceback
        traceback.print_exc()
        traceback.print_exc(file=sys.stderr)
        raise

